"""
BrainSAIT OCR Complete - Premium Dark Edition
AI-Powered OCR with Advanced Table Extraction
Futuristic Design | Enterprise Features | Excel Export
"""

import streamlit as st
import fitz  # PyMuPDF
import pytesseract
from PIL import Image
import pandas as pd
import io
import json
from datetime import datetime
import re
from pathlib import Path
import hashlib
import sqlite3
from typing import List, Dict, Optional, Tuple
import base64
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Page configuration
st.set_page_config(
    page_title="BrainSAIT OCR • Premium AI Platform",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Premium Dark Theme CSS
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    * {
        font-family: 'Inter', sans-serif;
    }
    
    /* Dark Background */
    .stApp {
        background: linear-gradient(135deg, #0a0e27 0%, #1a1d35 100%);
    }
    
    /* Premium Header */
    .premium-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 2.5rem;
        border-radius: 20px;
        text-align: center;
        margin-bottom: 2rem;
        box-shadow: 0 20px 60px rgba(102, 126, 234, 0.3);
        position: relative;
        overflow: hidden;
    }
    
    .premium-header::before {
        content: '';
        position: absolute;
        top: -50%;
        right: -50%;
        width: 200%;
        height: 200%;
        background: radial-gradient(circle, rgba(255,255,255,0.1) 0%, transparent 70%);
        animation: pulse 4s infinite;
    }
    
    @keyframes pulse {
        0%, 100% { transform: scale(1); opacity: 0.5; }
        50% { transform: scale(1.1); opacity: 0.8; }
    }
    
    .premium-title {
        font-size: 3rem;
        font-weight: 700;
        color: #ffffff;
        margin: 0;
        text-shadow: 0 4px 20px rgba(0,0,0,0.3);
        position: relative;
        z-index: 1;
    }
    
    .premium-subtitle {
        font-size: 1.2rem;
        color: rgba(255,255,255,0.9);
        margin-top: 0.5rem;
        position: relative;
        z-index: 1;
    }
    
    .premium-badge {
        display: inline-block;
        background: rgba(255,255,255,0.2);
        padding: 0.5rem 1.5rem;
        border-radius: 50px;
        color: white;
        font-weight: 600;
        margin-top: 1rem;
        backdrop-filter: blur(10px);
        border: 1px solid rgba(255,255,255,0.3);
        position: relative;
        z-index: 1;
    }
    
    /* Cards */
    .premium-card {
        background: rgba(255, 255, 255, 0.05);
        backdrop-filter: blur(20px);
        border: 1px solid rgba(255, 255, 255, 0.1);
        border-radius: 16px;
        padding: 2rem;
        margin: 1rem 0;
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3);
        transition: all 0.3s ease;
    }
    
    .premium-card:hover {
        transform: translateY(-5px);
        box-shadow: 0 12px 48px rgba(102, 126, 234, 0.4);
        border-color: rgba(102, 126, 234, 0.5);
    }
    
    /* Metrics */
    .metric-card {
        background: linear-gradient(135deg, rgba(102, 126, 234, 0.1) 0%, rgba(118, 75, 162, 0.1) 100%);
        border: 1px solid rgba(102, 126, 234, 0.3);
        border-radius: 12px;
        padding: 1.5rem;
        text-align: center;
        transition: all 0.3s ease;
    }
    
    .metric-card:hover {
        transform: scale(1.05);
        box-shadow: 0 8px 24px rgba(102, 126, 234, 0.3);
    }
    
    .metric-value {
        font-size: 2.5rem;
        font-weight: 700;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin: 0;
    }
    
    .metric-label {
        font-size: 0.9rem;
        color: rgba(255, 255, 255, 0.9);
        text-transform: uppercase;
        letter-spacing: 1px;
        margin-top: 0.5rem;
    }
    
    /* Buttons */
    .stButton>button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 12px;
        padding: 0.75rem 2rem;
        font-weight: 600;
        font-size: 1rem;
        transition: all 0.3s ease;
        box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);
    }
    
    .stButton>button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 25px rgba(102, 126, 234, 0.6);
    }
    
    /* File Uploader */
    .uploadedFile {
        background: rgba(255, 255, 255, 0.05);
        border: 2px dashed rgba(102, 126, 234, 0.5);
        border-radius: 12px;
        padding: 2rem;
        transition: all 0.3s ease;
    }
    
    .uploadedFile:hover {
        border-color: rgba(102, 126, 234, 0.8);
        background: rgba(102, 126, 234, 0.1);
    }
    
    /* Tabs */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background: rgba(255, 255, 255, 0.05);
        border-radius: 12px;
        padding: 0.5rem;
    }
    
    .stTabs [data-baseweb="tab"] {
        background: transparent;
        border-radius: 8px;
        color: rgba(255, 255, 255, 0.9);
        font-weight: 600;
        transition: all 0.3s ease;
    }
    
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
    }
    
    /* Progress Bar */
    .stProgress > div > div > div > div {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
    }
    
    /* Success/Info/Warning */
    .success-box {
        background: rgba(16, 185, 129, 0.1);
        border-left: 4px solid #10b981;
        border-radius: 8px;
        padding: 1rem;
        margin: 1rem 0;
        color: #10b981;
    }
    
    .info-box {
        background: rgba(59, 130, 246, 0.1);
        border-left: 4px solid #3b82f6;
        border-radius: 8px;
        padding: 1rem;
        margin: 1rem 0;
        color: #3b82f6;
    }
    
    .warning-box {
        background: rgba(245, 158, 11, 0.1);
        border-left: 4px solid #f59e0b;
        border-radius: 8px;
        padding: 1rem;
        margin: 1rem 0;
        color: #f59e0b;
    }
    
    /* Sidebar */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, rgba(10, 14, 39, 0.95) 0%, rgba(26, 29, 53, 0.95) 100%);
        backdrop-filter: blur(20px);
        border-right: 1px solid rgba(102, 126, 234, 0.2);
    }
    
    [data-testid="stSidebar"] .block-container {
        padding-top: 2rem;
    }
    
    /* DataFrames */
    .dataframe {
        background: rgba(255, 255, 255, 0.05);
        border-radius: 8px;
        overflow: hidden;
    }
    
    .dataframe thead tr th {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        font-weight: 600;
        padding: 1rem;
    }
    
    .dataframe tbody tr:nth-child(even) {
        background: rgba(255, 255, 255, 0.02);
    }
    
    .dataframe tbody tr:hover {
        background: rgba(102, 126, 234, 0.1);
    }
    
    /* Text Areas */
    .stTextArea textarea {
        background: rgba(255, 255, 255, 0.05);
        border: 1px solid rgba(255, 255, 255, 0.1);
        border-radius: 8px;
        color: white;
    }
    
    /* Download Buttons */
    .stDownloadButton>button {
        background: linear-gradient(135deg, #10b981 0%, #059669 100%);
        width: 100%;
    }
    
    /* Footer */
    .premium-footer {
        background: rgba(255, 255, 255, 0.05);
        border-top: 1px solid rgba(255, 255, 255, 0.1);
        padding: 2rem;
        margin-top: 3rem;
        border-radius: 16px;
        text-align: center;
    }
    
    /* Glow Effect */
    .glow {
        animation: glow 2s ease-in-out infinite;
    }
    
    @keyframes glow {
        0%, 100% { filter: drop-shadow(0 0 5px rgba(102, 126, 234, 0.5)); }
        50% { filter: drop-shadow(0 0 20px rgba(102, 126, 234, 0.8)); }
    }
    
    /* Loading Spinner */
    .stSpinner > div {
        border-top-color: #667eea;
    }
    
    /* Custom Scrollbar */
    ::-webkit-scrollbar {
        width: 10px;
        height: 10px;
    }
    
    ::-webkit-scrollbar-track {
        background: rgba(255, 255, 255, 0.05);
    }
    
    ::-webkit-scrollbar-thumb {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        border-radius: 5px;
    }
    
    ::-webkit-scrollbar-thumb:hover {
        background: linear-gradient(135deg, #764ba2 0%, #667eea 100%);
    }

    
    /* BRIGHT TEXT FIXES */
    .stApp, .stApp p, .stApp span, .stApp div, .stApp label, .stMarkdown {
        color: #FFFFFF !important;
    }
    
    h1, h2, h3, h4, h5, h6 {
        color: #FFFFFF !important;
    }
    
    .stTextInput input, .stTextArea textarea, .stSelectbox select {
        color: #FFFFFF !important;
        background: rgba(255, 255, 255, 0.1) !important;
    }
    
    .stTextInput label, .stSelectbox label, .stCheckbox label {
        color: #FFFFFF !important;
    }
    
    .stMetric label {
        color: #FFFFFF !important;
    }
    
    .stMetric .metric-value {
        color: #667eea !important;
    }

    
    /* SMART CONTRAST - Dark text on light backgrounds */
    .stButton>button {
        color: #FFFFFF !important;  /* White text on gradient buttons */
    }
    
    .premium-header, .premium-header * {
        color: #FFFFFF !important;  /* White text on gradient header */
    }
    
    .success-box, .success-box * {
        color: #10b981 !important;  /* Green text on success boxes */
    }
    
    .info-box, .info-box * {
        color: #60a5fa !important;  /* Bright blue on info boxes */
    }
    
    .warning-box, .warning-box * {
        color: #fbbf24 !important;  /* Bright yellow on warning boxes */
    }
    
    .dataframe thead tr th {
        color: #FFFFFF !important;  /* White text on table headers */
    }
    
    .dataframe tbody tr td {
        color: #1a1d35 !important;  /* Dark text on table cells (light bg) */
    }
    
    .stDownloadButton>button {
        color: #FFFFFF !important;  /* White text on download buttons */
    }
    
    /* Input fields - dark text when focused/filled */
    .stTextInput input:focus, .stTextArea textarea:focus {
        color: #1a1d35 !important;
        background: rgba(255, 255, 255, 0.9) !important;
    }
    
    /* Select boxes with light dropdown */
    .stSelectbox div[data-baseweb="select"] > div {
        color: #FFFFFF !important;
    }
    
    /* Expander headers */
    .streamlit-expanderHeader {
        color: #FFFFFF !important;
    }
    
    /* Metric cards - keep gradient text visible */
    .metric-value {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text !important;
        -webkit-text-fill-color: transparent !important;
    }
    
    .metric-label {
        color: rgba(255, 255, 255, 0.9) !important;
    }


    
    /* SIDEBAR - All text bright white */
    [data-testid="stSidebar"] * {
        color: #FFFFFF !important;
    }
    
    [data-testid="stSidebar"] h1, 
    [data-testid="stSidebar"] h2, 
    [data-testid="stSidebar"] h3,
    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] span,
    [data-testid="stSidebar"] div {
        color: #FFFFFF !important;
    }
    
    /* Sidebar select boxes */
    [data-testid="stSidebar"] .stSelectbox label,
    [data-testid="stSidebar"] .stSelectbox div,
    [data-testid="stSidebar"] .stSelectbox span {
        color: #FFFFFF !important;
    }
    
    /* Sidebar checkboxes */
    [data-testid="stSidebar"] .stCheckbox label,
    [data-testid="stSidebar"] .stCheckbox span {
        color: #FFFFFF !important;
    }
    
    /* FILE UPLOADER - All text bright white */
    .stFileUploader * {
        color: #FFFFFF !important;
    }
    
    .stFileUploader label,
    .stFileUploader span,
    .stFileUploader p,
    .stFileUploader div,
    .stFileUploader small {
        color: #FFFFFF !important;
    }
    
    /* File uploader drag area */
    [data-testid="stFileUploadDropzone"] *,
    [data-testid="stFileUploadDropzone"] label,
    [data-testid="stFileUploadDropzone"] span {
        color: #FFFFFF !important;
    }
    
    /* Upload instructions */
    .uploadedFileName {
        color: #FFFFFF !important;
    }
    
    /* Main content headings */
    .stApp h1, .stApp h2, .stApp h3, .stApp h4, .stApp h5, .stApp h6 {
        color: #FFFFFF !important;
    }
    
    /* Form labels */
    label {
        color: #FFFFFF !important;
    }
    
    /* Small text / help text */
    small, .stApp small {
        color: rgba(255, 255, 255, 0.8) !important;
    }

</style>
""", unsafe_allow_html=True)

# Initialize session state
if 'processing_history' not in st.session_state:
    st.session_state.processing_history = []
if 'current_results' not in st.session_state:
    st.session_state.current_results = None

class AdvancedTableDetector:
    """AI-Powered Table Detection and Extraction"""
    
    @staticmethod
    def detect_table_structure(text: str) -> List[Dict]:
        """Advanced table detection with AI-powered analysis"""
        tables = []
        lines = text.split('\n')
        
        current_table = []
        table_active = False
        min_columns = 2
        
        for i, line in enumerate(lines):
            line = line.strip()
            if not line:
                if table_active and len(current_table) >= 2:
                    tables.append(AdvancedTableDetector._process_table(current_table))
                table_active = False
                current_table = []
                continue
            
            # Detect columns by multiple separators
            cells = re.split(r'\s{2,}|\t|\|', line)
            cells = [c.strip() for c in cells if c.strip()]
            
            if len(cells) >= min_columns:
                if not table_active:
                    table_active = True
                current_table.append(cells)
            else:
                if table_active and len(current_table) >= 2:
                    tables.append(AdvancedTableDetector._process_table(current_table))
                table_active = False
                current_table = []
        
        # Add last table if exists
        if table_active and len(current_table) >= 2:
            tables.append(AdvancedTableDetector._process_table(current_table))
        
        return tables
    
    @staticmethod
    def _process_table(raw_table: List[List[str]]) -> Dict:
        """Process raw table data into structured format"""
        # Find max columns
        max_cols = max(len(row) for row in raw_table)
        
        # Normalize all rows to same column count
        normalized_table = []
        for row in raw_table:
            if len(row) < max_cols:
                row.extend([''] * (max_cols - len(row)))
            normalized_table.append(row[:max_cols])
        
        # Detect if first row is header (contains non-numeric values)
        has_header = False
        if normalized_table:
            first_row = normalized_table[0]
            numeric_count = sum(1 for cell in first_row if re.match(r'^[\d\.,]+$', cell))
            has_header = numeric_count < len(first_row) / 2
        
        return {
            'rows': len(normalized_table),
            'columns': max_cols,
            'has_header': has_header,
            'header': normalized_table[0] if has_header else [f'Column {i+1}' for i in range(max_cols)],
            'data': normalized_table[1:] if has_header else normalized_table,
            'full_data': normalized_table
        }
    
    @staticmethod
    def export_to_excel(tables: List[Dict], filename: str) -> bytes:
        """Export tables to professionally formatted Excel file"""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Styling
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='667EEA', end_color='764BA2', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Arial', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        border_style = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        for idx, table in enumerate(tables, 1):
            ws = wb.create_sheet(title=f'Table {idx}')
            
            # Write header
            for col_idx, header in enumerate(table['header'], 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_style
            
            # Write data
            for row_idx, row_data in enumerate(table['data'], 2):
                for col_idx, cell_value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    cell.font = cell_font
                    cell.alignment = cell_alignment
                    cell.border = border_style
                    
                    # Auto-detect numeric values
                    if re.match(r'^[\d\.,]+$', str(cell_value)):
                        try:
                            cell.value = float(cell_value.replace(',', ''))
                            cell.number_format = '#,##0.00'
                        except:
                            pass
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

class OCRProcessor:
    """Enhanced OCR processing engine"""
    
    def __init__(self):
        self.supported_formats = ['pdf', 'png', 'jpg', 'jpeg', 'webp', 'bmp', 'tiff']
        self.table_detector = AdvancedTableDetector()
    
    def calculate_file_hash(self, file_bytes: bytes) -> str:
        return hashlib.sha256(file_bytes).hexdigest()
    
    def extract_text_from_image(self, image: Image.Image, lang: str = 'eng+ara') -> str:
        try:
            # Enhanced preprocessing for better OCR
            image = image.convert('L')  # Grayscale
            # Enhance contrast
            image_array = np.array(image)
            image_array = np.clip(image_array * 1.2, 0, 255).astype(np.uint8)
            image = Image.fromarray(image_array)
            
            text = pytesseract.image_to_string(
                image, 
                lang=lang,
                config='--psm 3 --oem 3'  # Best mode for documents
            )
            return text
        except Exception as e:
            st.error(f"OCR Error: {str(e)}")
            return ""
    
    def extract_from_pdf(self, pdf_bytes: bytes, lang: str = 'eng+ara', 
                        use_ocr: bool = True, progress_callback=None) -> Dict:
        results = {
            'pages': [],
            'total_text': '',
            'metadata': {},
            'tables': [],
            'page_count': 0
        }
        
        try:
            pdf = fitz.open(stream=pdf_bytes, filetype="pdf")
            results['page_count'] = pdf.page_count
            results['metadata'] = pdf.metadata
            
            for page_num in range(pdf.page_count):
                if progress_callback:
                    progress_callback(page_num + 1, pdf.page_count)
                
                page = pdf[page_num]
                text = page.get_text()
                
                if use_ocr and len(text.strip()) < 50:
                    pix = page.get_pixmap(matrix=fitz.Matrix(2.5, 2.5))  # Higher quality
                    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    text = self.extract_text_from_image(img, lang)
                
                # Advanced table detection
                tables = self.table_detector.detect_table_structure(text)
                
                page_data = {
                    'page_number': page_num + 1,
                    'text': text,
                    'char_count': len(text),
                    'word_count': len(text.split()),
                    'tables': tables,
                    'table_count': len(tables)
                }
                
                results['pages'].append(page_data)
                results['total_text'] += f"\n\n=== Page {page_num + 1} ===\n\n{text}"
                results['tables'].extend(tables)
            
            pdf.close()
            
        except Exception as e:
            st.error(f"PDF Processing Error: {str(e)}")
        
        return results

def main():
    # Premium Header
    st.markdown("""
    <div class="premium-header">
        <h1 class="premium-title">⚡ BrainSAIT OCR</h1>
        <p class="premium-subtitle">AI-Powered Document Intelligence • Premium Edition</p>
        <div class="premium-badge">🚀 Powered by Tesseract 5.0 + Advanced AI</div>
    </div>
    """, unsafe_allow_html=True)
    
    processor = OCRProcessor()
    
    # Sidebar
    with st.sidebar:
        st.markdown("### ⚙️ Advanced Settings")
        
        languages = {
            '🌍 English + Arabic': 'eng+ara',
            '🇬🇧 English Only': 'eng',
            '🇸🇦 Arabic Only': 'ara',
            '🇫🇷 French + Arabic': 'fra+ara',
            '🇪🇸 Spanish + English': 'spa+eng',
        }
        
        selected_lang = st.selectbox(
            "OCR Language",
            options=list(languages.keys()),
            index=0
        )
        lang_code = languages[selected_lang]
        
        enable_ocr = st.checkbox("🔍 Enable OCR for Scanned PDFs", value=True)
        extract_tables = st.checkbox("📊 AI Table Extraction", value=True)
        
        st.markdown("---")
        st.markdown("### 💎 Premium Features")
        st.markdown("""
        ✅ Advanced AI table detection  
        ✅ Professional Excel export  
        ✅ Multi-format support  
        ✅ Intelligent text extraction  
        ✅ Premium dark theme  
        """)
    
    # Main Upload Area
    st.markdown('<div class="premium-card">', unsafe_allow_html=True)
    
    col1, col2 = st.columns([3, 1])
    
    with col1:
        st.markdown("### 📤 Upload Your Document")
        uploaded_file = st.file_uploader(
            "Drop your PDF or image here",
            type=processor.supported_formats,
            help="Supports: PDF, PNG, JPG, JPEG, WEBP, BMP, TIFF"
        )
    
    with col2:
        if uploaded_file:
            st.success("✅ File Ready")
            st.metric("Size", f"{uploaded_file.size / 1024:.1f} KB")
            file_ext = uploaded_file.name.split('.')[-1].upper()
            st.metric("Type", file_ext)
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    if uploaded_file is not None:
        if st.button("🚀 Start AI Processing", type="primary", use_container_width=True):
            file_bytes = uploaded_file.read()
            file_hash = processor.calculate_file_hash(file_bytes)
            file_ext = uploaded_file.name.split('.')[-1].lower()
            
            start_time = datetime.now()
            
            with st.spinner('⚡ AI Processing in progress...'):
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                def update_progress(current, total):
                    progress = current / total
                    progress_bar.progress(progress)
                    status_text.markdown(f'<div class="info-box">Processing page {current}/{total}...</div>', unsafe_allow_html=True)
                
                if file_ext == 'pdf':
                    results = processor.extract_from_pdf(
                        file_bytes, 
                        lang=lang_code, 
                        use_ocr=enable_ocr,
                        progress_callback=update_progress
                    )
                else:
                    img = Image.open(io.BytesIO(file_bytes))
                    text = processor.extract_text_from_image(img, lang_code)
                    tables = processor.table_detector.detect_table_structure(text)
                    results = {
                        'pages': [{'page_number': 1, 'text': text, 'char_count': len(text),
                                  'word_count': len(text.split()), 'tables': tables, 'table_count': len(tables)}],
                        'total_text': text,
                        'tables': tables
                    }
                
                processing_time = (datetime.now() - start_time).total_seconds()
                progress_bar.empty()
                status_text.empty()
            
            st.session_state.current_results = results
            
            st.markdown(f'<div class="success-box">✅ Processing complete in {processing_time:.2f} seconds!</div>', unsafe_allow_html=True)
            
            # Premium Metrics
            st.markdown("### 📊 Document Insights")
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.markdown(f'''
                <div class="metric-card">
                    <div class="metric-value">{len(results.get("pages", []))}</div>
                    <div class="metric-label">Pages</div>
                </div>
                ''', unsafe_allow_html=True)
            
            with col2:
                total_chars = len(results.get('total_text', ''))
                st.markdown(f'''
                <div class="metric-card">
                    <div class="metric-value">{total_chars:,}</div>
                    <div class="metric-label">Characters</div>
                </div>
                ''', unsafe_allow_html=True)
            
            with col3:
                total_words = sum(p['word_count'] for p in results.get('pages', []))
                st.markdown(f'''
                <div class="metric-card">
                    <div class="metric-value">{total_words:,}</div>
                    <div class="metric-label">Words</div>
                </div>
                ''', unsafe_allow_html=True)
            
            with col4:
                tables_count = len(results.get('tables', []))
                st.markdown(f'''
                <div class="metric-card">
                    <div class="metric-value">{tables_count}</div>
                    <div class="metric-label">Tables</div>
                </div>
                ''', unsafe_allow_html=True)
            
            # Results Tabs
            st.markdown("### 🎯 Results")
            tab1, tab2, tab3, tab4 = st.tabs([
                "📝 Extracted Text",
                "📊 AI Tables",
                "🔍 Smart Search",
                "💾 Export Options"
            ])
            
            with tab1:
                st.markdown('<div class="premium-card">', unsafe_allow_html=True)
                st.text_area(
                    "Full Extracted Text",
                    value=results.get('total_text', ''),
                    height=400
                )
                st.markdown('</div>', unsafe_allow_html=True)
            
            with tab2:
                if results.get('tables'):
                    st.markdown(f'<div class="success-box">✅ Found {len(results["tables"])} tables with AI detection</div>', unsafe_allow_html=True)
                    
                    for idx, table in enumerate(results['tables'], 1):
                        with st.expander(f"📊 Table {idx} - {table['rows']} rows × {table['columns']} columns", expanded=True):
                            # Create DataFrame
                            df = pd.DataFrame(table['data'], columns=table['header'])
                            st.dataframe(df, use_container_width=True)
                            
                            # Download individual table
                            csv = df.to_csv(index=False)
                            st.download_button(
                                label=f"📥 Download Table {idx} as CSV",
                                data=csv,
                                file_name=f"{uploaded_file.name}_table_{idx}.csv",
                                mime="text/csv"
                            )
                else:
                    st.markdown('<div class="info-box">ℹ️ No tables detected in this document</div>', unsafe_allow_html=True)
            
            with tab3:
                st.markdown('<div class="premium-card">', unsafe_allow_html=True)
                search_term = st.text_input("🔍 Search in document")
                
                if search_term:
                    matches = []
                    for page in results.get('pages', []):
                        lines = page['text'].split('\n')
                        for line in lines:
                            if search_term.lower() in line.lower():
                                matches.append({
                                    'Page': page['page_number'],
                                    'Text': line.strip()[:200]
                                })
                    
                    if matches:
                        st.success(f"✅ Found {len(matches)} matches")
                        df_matches = pd.DataFrame(matches)
                        st.dataframe(df_matches, use_container_width=True)
                    else:
                        st.warning(f"No matches found for '{search_term}'")
                st.markdown('</div>', unsafe_allow_html=True)
            
            with tab4:
                st.markdown("### 💾 Professional Export Options")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown('<div class="premium-card">', unsafe_allow_html=True)
                    st.markdown("#### 📄 Text Formats")
                    
                    full_text = results.get('total_text', '')
                    
                    # TXT Export
                    st.download_button(
                        label="📄 Download Full Text (.txt)",
                        data=full_text,
                        file_name=f"{uploaded_file.name}_extracted.txt",
                        mime="text/plain",
                        use_container_width=True
                    )
                    
                    # JSON Export
                    json_data = json.dumps(results, indent=2, ensure_ascii=False)
                    st.download_button(
                        label="📊 Download Analysis (.json)",
                        data=json_data,
                        file_name=f"{uploaded_file.name}_analysis.json",
                        mime="application/json",
                        use_container_width=True
                    )
                    st.markdown('</div>', unsafe_allow_html=True)
                
                with col2:
                    st.markdown('<div class="premium-card">', unsafe_allow_html=True)
                    st.markdown("#### 📊 Excel Export")
                    
                    if results.get('tables'):
                        # Generate base filename
                        base_name = uploaded_file.name.rsplit('.', 1)[0]
                        excel_filename = f"{base_name}_tables.xlsx"
                        
                        excel_data = processor.table_detector.export_to_excel(
                            results['tables'],
                            excel_filename
                        )
                        
                        st.download_button(
                            label=f"📊 Download All Tables (.xlsx)",
                            data=excel_data,
                            file_name=excel_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                        
                        st.success(f"✅ {len(results['tables'])} tables ready for Excel export")
                        st.info(f"📁 File: {excel_filename}")
                    else:
                        st.warning("No tables available for Excel export")
                    
                    st.markdown('</div>', unsafe_allow_html=True)
    
    # Premium Footer
    st.markdown("""
    <div class="premium-footer">
        <p style="color: rgba(255,255,255,0.8); font-size: 0.9rem;">
            <strong>BrainSAIT OCR</strong> • Premium AI Platform<br>
            Powered by Tesseract 5.0 + Advanced AI • Built with Streamlit<br>
            © 2026 Dr. Mohammed Al-Fadil | BrainSAIT
        </p>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
